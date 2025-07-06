import asyncio
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from doc_parser.config import AppConfig as Settings
from doc_parser.parsers.excel.parser import ExcelParser


@pytest.fixture()
def make_sample_excel(tmp_path):  # noqa: D401
    """Return path to a generated .xlsx workbook with small data."""

    def _factory() -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = "Name"
        ws["B1"].value = "Age"
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])

        file_path = tmp_path / "sample.xlsx"
        wb.save(file_path)
        return file_path

    return _factory


@pytest.mark.asyncio
async def test_excel_parser_markdown(make_sample_excel):
    xlsx_path = make_sample_excel()

    settings = Settings(output_format="markdown")
    parser = ExcelParser(settings)
    result = await parser.parse(xlsx_path)

    assert result.content.startswith("# Sheet: Sheet1")
    # Table header row should be present
    assert "| Name | Age |" in result.content
    # Metadata checks
    assert result.metadata["sheet_count"] == 1
    assert result.format == "markdown"


@pytest.mark.asyncio
async def test_excel_parser_json(make_sample_excel):
    xlsx_path = make_sample_excel()
    settings = Settings(output_format="json")
    parser = ExcelParser(settings)
    result = await parser.parse(xlsx_path)

    # JSON string should include sheet name and data keys
    assert "\"Sheet1\"" in result.content
    assert "\"data\"" in result.content


def test_excel_validate_input_neg(tmp_path):
    fake = tmp_path / "file.txt"
    fake.write_text("x")
    parser = ExcelParser(Settings())
    assert asyncio.run(parser.validate_input(fake)) is False 
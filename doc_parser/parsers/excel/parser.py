"""Excel parser implementation.

This module provides a parser for Microsoft Excel files (.xlsx, .xls, .xlsm). It
supports extraction of sheet data into Markdown or JSON, optional formulas and
formatting, and metadata aggregation (sheet names, counts).

Examples:
>>> from pathlib import Path
>>> import asyncio
>>> from doc_parser.parsers.excel.parser import ExcelParser
>>> from doc_parser.core.settings import Settings
>>> settings = Settings(output_format="json", parser_settings={"excel": {"include_formulas": True}})
>>> parser = ExcelParser(settings)
>>> result = asyncio.run(parser.parse(Path("example.xlsx")))
>>> assert "sheets" in result.metadata
"""

import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd

from doc_parser.config import AppConfig
from doc_parser.core.base import BaseParser
from doc_parser.core.error_policy import EXPECTED_EXCEPTIONS
from doc_parser.utils.mixins import DataFrameMarkdownMixin

if TYPE_CHECKING:  # pragma: no cover
    from pydantic import BaseModel


@AppConfig.register("excel", [".xlsx", ".xls", ".xlsm"])
class ExcelParser(DataFrameMarkdownMixin, BaseParser):
    """Parser for Excel files (.xlsx, .xls, .xlsm).

    Args:
        config (Settings): Global settings, with optional Excel-specific keys:
            - include_formulas (bool): Whether to extract cell formulas (default False).
            - include_formatting (bool): Whether to include cell formatting (default False).
            - sheet_names (List[str] | None): Specific sheets to parse; None for all (default).

    Attributes:
        include_formulas (bool): If True, extract formulas.
        include_formatting (bool): If True, include formatting.
        sheet_names (List[str] | None): Sheets to process.

    Examples:
        >>> from pathlib import Path
        >>> import asyncio
        >>> from doc_parser.parsers.excel.parser import ExcelParser
        >>> from doc_parser.core.settings import Settings
        >>> settings = Settings(output_format="markdown", parser_settings={"excel": {"include_formulas": True}})
        >>> parser = ExcelParser(settings)
        >>> result = asyncio.run(parser.parse(Path("data.xlsx")))
        >>> print(result.format)
        'markdown'
    """

    def __init__(self, config: AppConfig):
        """Initialize Excel parser."""
        super().__init__(config)

        # Get Excel-specific settings (typed model)
        excel_cfg = config.parsers.excel

        self.include_formulas = (
            excel_cfg.include_formulas if excel_cfg.include_formulas is not None else False
        )
        self.include_formatting = (
            excel_cfg.include_formatting if excel_cfg.include_formatting is not None else False
        )
        self.sheet_names = excel_cfg.sheet_names  # None implies all sheets

    async def validate_input(self, input_path: Path) -> bool:
        """Validate whether the input path points to a readable Excel file.

        Args:
            input_path (Path): Path to the Excel file.

        Returns:
            bool: True if file exists, has supported extension, and is readable via pandas.

        Example:
            >>> import asyncio
            >>> from pathlib import Path
            >>> parser = ExcelParser(Settings())
            >>> valid = asyncio.run(parser.validate_input(Path("file.xlsx")))
            >>> print(valid)
        """
        if not self._has_supported_extension(input_path):
            return False
        try:
            # Try to open with pandas to validate
            pd.ExcelFile(input_path)
        except (InvalidFileException, ValueError, OSError):
            return False
        return True

    # ------------------------------------------------------------------
    # BaseStructuredParser hooks
    # ------------------------------------------------------------------

    async def _open_document(self, input_path: Path, *, options: "BaseModel | None" = None) -> Path:
        """Return *input_path* so downstream helpers can open via pandas as needed."""
        _ = options
        return input_path

    def _extra_metadata(self, input_path: Any) -> dict[str, Any]:
        """Return sheet names and count for quick metadata lookup."""
        logger = logging.getLogger(__name__)
        try:
            excel_file = pd.ExcelFile(input_path)
            return {
                "sheets": excel_file.sheet_names,
                "sheet_count": len(excel_file.sheet_names),
            }
        except EXPECTED_EXCEPTIONS as exc:
            logger.debug("Expected error reading Excel metadata for %s: %s", input_path, exc, exc_info=True)
            return {}

    async def _extract_as_markdown(self, document_obj: Path) -> str:
        """Extract Excel content as a Markdown string.

        Iterates over sheets, converts DataFrames to markdown tables, includes empty sheet markers,
        and appends formulas if configured.

        Args:
            document_obj (Path): Path to the Excel file.

        Returns:
            str: Combined Markdown content for all processed sheets.

        Example:
            >>> import asyncio
            >>> from pathlib import Path
            >>> parser = ExcelParser(Settings(parser_settings={"excel": {"include_formulas": True}}))
            >>> md = asyncio.run(parser._extract_as_markdown(Path("file.xlsx")))
            >>> assert "# Sheet:" in md
        """
        excel_file = pd.ExcelFile(document_obj)
        content_parts = []

        # Determine which sheets to process
        if self.sheet_names is not None:
            sheets_to_process: list[str] = [s for s in self.sheet_names if s in excel_file.sheet_names]
        else:
            sheets_to_process = [str(name) for name in excel_file.sheet_names]

        for sheet_name in sheets_to_process:
            # Add sheet header
            content_parts.append(f"# Sheet: {sheet_name}\n")

            # Read sheet data
            df = pd.read_excel(document_obj, sheet_name=sheet_name, header=None)

            # Convert to markdown table
            if not df.empty:
                markdown_table = self._dataframe_to_markdown(df)
                content_parts.append(markdown_table)
            else:
                content_parts.append("*Empty sheet*")

            # Add formulas if requested
            if self.include_formulas:
                formulas = await self._extract_formulas(document_obj, sheet_name)
                if formulas:
                    content_parts.append("\n## Formulas\n")
                    for cell, formula in formulas.items():
                        content_parts.append(f"- {cell}: `{formula}`")

            content_parts.append("\n")

        return "\n".join(content_parts)

    async def _extract_as_json(self, document_obj: Path) -> str:
        """Extract Excel content as a JSON string.

        Serializes each sheet into an object containing data records, column names, shape,
        and optional formulas.

        Args:
            document_obj (Path): Path to the Excel file.

        Returns:
            str: JSON-formatted string of sheet data.

        Example:
            >>> import asyncio, json
            >>> from pathlib import Path
            >>> parser = ExcelParser(Settings())
            >>> js = asyncio.run(parser._extract_as_json(Path("file.xlsx")))
            >>> data = json.loads(js)
            >>> assert isinstance(data, dict)
        """
        import json

        excel_file = pd.ExcelFile(document_obj)
        data = {}

        # Determine which sheets to process
        if self.sheet_names is not None:
            sheets_to_process: list[str] = [s for s in self.sheet_names if s in excel_file.sheet_names]
        else:
            sheets_to_process = [str(name) for name in excel_file.sheet_names]

        for sheet_name in sheets_to_process:
            df = pd.read_excel(document_obj, sheet_name=sheet_name)

            # Convert to dict
            sheet_data = {
                "data": df.to_dict(orient="records"),
                "columns": df.columns.tolist(),
                "shape": df.shape,
            }

            # Add formulas if requested
            if self.include_formulas:
                sheet_data["formulas"] = await self._extract_formulas(document_obj, sheet_name)

            data[sheet_name] = sheet_data

        return json.dumps(data, indent=2, default=str)

    async def _extract_formulas(self, input_path: Path, sheet_name: str) -> dict[str, str]:
        """Extract cell formulas from a specific sheet of an Excel file.

        Args:
            input_path (Path): Path to the Excel file.
            sheet_name (str): Name of the sheet to extract formulas from.

        Returns:
            Dict[str, str]: Mapping of cell references to formula strings.

        Example:
            >>> import asyncio
            >>> from pathlib import Path
            >>> parser = ExcelParser(Settings(parser_settings={"excel": {"include_formulas": True}}))
            >>> formulas = asyncio.run(parser._extract_formulas(Path("file.xlsx"), "Sheet1"))
            >>> assert isinstance(formulas, dict)
        """
        formulas: dict[str, str] = {}
        try:
            wb = openpyxl.load_workbook(str(input_path), data_only=False)
            try:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                            formulas[cell_ref] = cell.value
            finally:
                wb.close()
        except (InvalidFileException, KeyError, OSError):
            # Failed to load workbook or sheet not found
            return {}
        return formulas
